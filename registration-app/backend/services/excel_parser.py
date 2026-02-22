"""
Excel import/export service using openpyxl.

Template columns (case-insensitive):
  name | email | company | modules

modules: comma-separated list of bundle IDs, e.g. "app-development, istio-service-mesh"
Valid bundle IDs are loaded at runtime from courses.yaml.
"""

import io
import os
import yaml
from typing import Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

REQUIRED_COLUMNS = {"name", "email", "modules"}

COURSES_YAML = os.path.join(os.path.dirname(__file__), "..", "..", "..", "courses.yaml")


def _load_valid_bundle_ids() -> List[str]:
    path = os.path.abspath(COURSES_YAML)
    if not os.path.exists(path):
        path = "courses.yaml"
    with open(path) as f:
        data = yaml.safe_load(f)
    return list(data.get("bundles", {}).keys())


class ExcelParser:
    def parse(self, file_bytes: bytes):
        """
        Parse an .xlsx file and return (rows, errors).

        rows: list of dicts with keys: name, email, company, modules (list of str)
        errors: list of human-readable error strings for invalid rows
        """
        valid_bundles = _load_valid_bundle_ids()

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active

        # Read header row
        headers = [str(cell.value).strip().lower() for cell in ws[1] if cell.value]
        missing = REQUIRED_COLUMNS - set(headers)
        if missing:
            return [], [f"Missing required columns: {', '.join(sorted(missing))}"]

        rows = []
        errors = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            values = dict(zip(headers, [str(v).strip() if v else "" for v in row]))

            # Skip completely empty rows
            if not any(values.values()):
                continue

            if not values.get("name"):
                errors.append(f"Row {row_num}: missing name")
                continue
            if not values.get("email") or "@" not in values["email"]:
                errors.append(f"Row {row_num}: invalid or missing email")
                continue

            raw_modules = values.get("modules", "")
            module_list = [m.strip() for m in raw_modules.split(",") if m.strip()]
            if not module_list:
                errors.append(f"Row {row_num}: modules column is empty — enter at least one bundle ID")
                continue
            invalid = [m for m in module_list if m not in valid_bundles]
            if invalid:
                errors.append(f"Row {row_num}: unknown bundle(s): {invalid}. Valid: {valid_bundles}")
                continue

            rows.append({
                "name": values["name"],
                "email": values["email"].lower(),
                "company": values.get("company", ""),
                "modules": module_list,
            })

        return rows, errors

    def generate_template(self) -> bytes:
        """Generate a formatted Excel template for download."""
        valid_bundles = _load_valid_bundle_ids()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Participants"

        headers = ["name", "email", "company", "modules"]
        header_fill = PatternFill(start_color="4B00AA", end_color="4B00AA", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, name="Montserrat")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = 30

        # Example rows
        examples = [
            ["Alex Chen", "alex.chen@example.com", "Acme Corp", "app-development, istio-service-mesh"],
            ["Jordan Lee", "jordan.lee@example.com", "Acme Corp", "infra-fundamentals, cluster-operations, storage-and-dr"],
        ]
        for row_data in examples:
            ws.append(row_data)

        # Note row with valid bundle IDs
        ws.append([])
        note_text = "modules: comma-separated bundle IDs — valid values: " + ", ".join(valid_bundles)
        note = ws.cell(row=ws.max_row + 1, column=1, value=note_text)
        note.font = Font(italic=True, color="888888")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
